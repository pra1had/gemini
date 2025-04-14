import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from backend.models import Scenario

def generate_excel(scenario: Scenario) -> bytes:
    """
    Generates a standard Excel (.xlsx) file content as bytes from a Scenario object.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Scenario Details"

    # --- Metadata Section ---
    ws['A1'] = "Scenario Name:"
    ws['B1'] = scenario.scenario_name
    ws['A2'] = "Component:"
    ws['B2'] = scenario.component_name or "N/A"
    ws['A3'] = "Tags:"
    ws['B3'] = ", ".join(scenario.tags) if scenario.tags else "N/A"

    # Style metadata headers
    for cell in ['A1', 'A2', 'A3']:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal='right')

    # --- Flow Steps Section ---
    current_row = 5 # Start steps below metadata
    ws.cell(row=current_row, column=1, value="Scenario Flow Steps").font = Font(bold=True, size=14)
    current_row += 2

    if not scenario.flow_steps:
        ws.cell(row=current_row, column=1, value="(No steps defined)")
        current_row += 1
    else:
        # Header for the steps overview
        ws.cell(row=current_row, column=1, value="Step").font = Font(bold=True)
        ws.cell(row=current_row, column=2, value="Action Code").font = Font(bold=True)
        current_row += 1

        # List each step and its action code
        for i, step in enumerate(scenario.flow_steps):
            ws.cell(row=current_row + i, column=1, value=i + 1)
            ws.cell(row=current_row + i, column=2, value=step.action_code)

        current_row += len(scenario.flow_steps) + 2 # Add spacing before detailed data

        # --- Step Data Details Section ---
        ws.cell(row=current_row, column=1, value="Step Data Details").font = Font(bold=True, size=14)
        current_row += 2

        for i, step in enumerate(scenario.flow_steps):
            ws.cell(row=current_row, column=1, value=f"Step {i+1}: {step.action_code}").font = Font(bold=True)
            current_row += 1

            if not step.step_data:
                ws.cell(row=current_row, column=1, value="(No data for this step)")
                current_row += 2 # Add spacing
                continue

            # Create headers from keys of the first data row (assuming consistent keys)
            headers = list(step.step_data[0].keys())
            for col_idx, header in enumerate(headers):
                cell = ws.cell(row=current_row, column=col_idx + 1, value=header)
                cell.font = Font(italic=True)
                cell.alignment = Alignment(horizontal='left')

            current_row += 1

            # Add data rows
            for row_data in step.step_data:
                for col_idx, header in enumerate(headers):
                    ws.cell(row=current_row, column=col_idx + 1, value=str(row_data.get(header, ''))) # Ensure value is string
                current_row += 1

            current_row += 1 # Add spacing between steps

    # Adjust column widths (optional, but improves readability)
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 20 # Adjust width as needed

    # Save workbook to a bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
